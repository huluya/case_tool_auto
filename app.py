import os
import io
import json
import html
import shutil
import subprocess
import re
import posixpath
from zipfile import BadZipFile, ZipFile
from xml.etree import ElementTree as ET
from datetime import datetime

from flask import Flask, request, jsonify, send_file, render_template, session
from sqlalchemy import create_engine, text
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.cell_range import CellRange

import config
from models import (
    db, Project, Version, CustomColumn, TestCase, CaseImage, CaseMerge,
    Role, User, SYSTEM_COLUMNS, STATUS_LIST
)

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = config.SQLALCHEMY_DATABASE_URI
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB
app.secret_key = config.SECRET_KEY

db.init_app(app)


ROLE_READONLY = 0
ROLE_TEST = 1
ROLE_ADMIN = 2
ADMIN_ENDPOINTS = {
    'update_project', 'delete_project',
    'update_version', 'delete_version', 'copy_version',
    'backup_database',
}


def migrate_auth_schema():
    """为已有 roles 表补充数据库维护的权限字段。"""
    with db.engine.begin() as conn:
        for column, definition in (
            ('can_write', 'TINYINT(1) NULL DEFAULT NULL'),
            ('can_manage', 'TINYINT(1) NULL DEFAULT NULL'),
        ):
            try:
                conn.execute(text(f'ALTER TABLE roles ADD COLUMN {column} {definition}'))
            except Exception:
                pass  # 字段已存在


def initialize_auth_data():
    """创建权限表并补齐首次运行所需的角色和账号。"""
    try:
        db.create_all()
    except Exception:
        # 支持直接使用 flask run 的场景：数据库尚未创建时先创建库，
        # 再创建 roles/users 等应用表。
        create_database()
        db.create_all()
    migrate_auth_schema()
    role_seeds = (
        (ROLE_READONLY, '只读', '只能查看和导出数据', False, False),
        (ROLE_TEST, '测试', '可以编辑用例和执行结果', True, False),
        (ROLE_ADMIN, '管理员', '可以管理项目、版本、用例和备份', True, True),
    )
    roles = {}
    for status, name, description, can_write, can_manage in role_seeds:
        role = Role.query.filter_by(status=status).first()
        if not role:
            role = Role(
                status=status, name=name, description=description,
                can_write=can_write, can_manage=can_manage, is_active=True
            )
            db.session.add(role)
            db.session.flush()
        else:
            # 旧版本角色表没有权限字段时，迁移出的 NULL 才使用默认值；
            # 后续管理员在数据库中的修改不会被启动流程覆盖。
            if role.can_write is None:
                role.can_write = can_write
            if role.can_manage is None:
                role.can_manage = can_manage
        roles[status] = role

    user_seeds = (
        ('admin', '123456', ROLE_ADMIN),
        ('test', '123456', ROLE_TEST),
        ('readonly', '123456', ROLE_READONLY),
    )
    for username, password, role_status in user_seeds:
        user = User.query.filter_by(username=username).first()
        if not user:
            db.session.add(User(
                username=username,
                password_hash=generate_password_hash(password),
                role_id=roles[role_status].id,
                is_active=True,
            ))
        elif not user.role_id:
            user.role_id = roles[role_status].id
    db.session.commit()


def current_user():
    """返回当前会话账号；角色是否启用由数据库决定。"""
    user_id = session.get('user_id')
    if not user_id:
        return None
    return User.query.join(Role).filter(
        User.id == user_id,
        User.is_active.is_(True),
        Role.is_active.is_(True),
    ).first()


@app.before_request
def require_api_login():
    if not request.path.startswith('/api'):
        return None
    if request.method == 'OPTIONS':
        return None
    if request.path == '/api/auth/login' or request.endpoint == 'logout':
        return None
    user = current_user()
    if not user:
        return jsonify({'success': False, 'message': '请先登录'}), 401
    if not user.role.can_write and request.method != 'GET':
        return jsonify({'success': False, 'message': '只读账号不可操作'}), 403
    if request.endpoint in ADMIN_ENDPOINTS and not user.role.can_manage:
        return jsonify({'success': False, 'message': '当前账号没有管理员权限'}), 403
    return None


def create_database():
    """如果数据库不存在则创建"""
    engine = create_engine(
        f"mysql+pymysql://{config.DB_USER}:{config.DB_PASSWORD}@{config.DB_HOST}:{config.DB_PORT}",
        isolation_level="AUTOCOMMIT"
    )
    with engine.connect() as conn:
        conn.execute(text(f"CREATE DATABASE IF NOT EXISTS {config.DB_NAME} CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"))


def init_system_columns(project_id, version_id):
    """初始化指定版本的系统列，列配置严格按版本隔离。"""
    if not version_id:
        return
    columns = CustomColumn.query.filter_by(
        project_id=project_id, version_id=version_id
    ).all()
    if not columns:
        # 兼容尚未完成旧数据迁移的数据库：把历史项目级列归到当前版本。
        legacy = CustomColumn.query.filter_by(
            project_id=project_id, version_id=None
        ).all()
        if legacy:
            for column in legacy:
                column.version_id = version_id
            columns = legacy

    by_key = {column.key: column for column in columns}
    for idx, col in enumerate(SYSTEM_COLUMNS):
        exists = by_key.get(col['key'])
        if not exists:
            legacy = next((column for column in columns
                           if column.name == col['name'] and not column.is_system), None)
            if legacy:
                for case in TestCase.query.filter_by(
                        project_id=project_id, version_id=version_id).all():
                    custom = case.get_custom_fields()
                    if legacy.key in custom:
                        setattr(case, col['key'], custom.pop(legacy.key))
                        case.set_custom_fields(custom)
                legacy.name = col['name']
                legacy.key = col['key']
                legacy.is_system = True
                legacy.width = col['width']
                legacy.sort_order = idx
                by_key[col['key']] = legacy
            else:
                created = CustomColumn(
                    project_id=project_id,
                    version_id=version_id,
                    name=col['name'],
                    key=col['key'],
                    is_system=col['is_system'],
                    is_visible=True,
                    width=col['width'],
                    sort_order=idx
                )
                db.session.add(created)
    db.session.commit()


def query_version_columns(project_id, version_id):
    """返回某个版本的全部列，禁止跨版本读取列配置。"""
    return CustomColumn.query.filter_by(
        project_id=project_id, version_id=version_id
    ).order_by(CustomColumn.sort_order.asc(), CustomColumn.id.asc()).all()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/login', methods=['POST'])
@app.route('/api/auth/login', methods=['POST'])
def login():
    initialize_auth_data()
    data = request.json or {}
    username = (data.get('username') or '').strip()
    password = data.get('password', '')
    user = User.query.filter_by(username=username, is_active=True).first()
    if user and user.role and user.role.is_active and check_password_hash(user.password_hash, password):
        session.clear()
        session['user_id'] = user.id
        role_data = user.to_dict()
        return jsonify({
            'success': True,
            'data': role_data
        })
    return jsonify({'success': False, 'message': '账号或密码错误'}), 401


@app.route('/api/auth/me', methods=['GET'])
def auth_me():
    user = current_user()
    return jsonify({'success': True, 'data': user.to_dict()})


@app.route('/api/auth/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})


# ------------------- 项目 -------------------
@app.route('/api/projects', methods=['GET'])
def get_projects():
    projects = Project.query.order_by(Project.created_at.desc()).all()
    return jsonify({'success': True, 'data': [p.to_dict() for p in projects]})


@app.route('/api/projects', methods=['POST'])
def create_project():
    data = request.json or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'message': '项目名称不能为空'}), 400
    if Project.query.filter_by(name=name).first():
        return jsonify({'success': False, 'message': '项目名称已存在'}), 400

    project = Project(name=name, description=data.get('description', ''))
    db.session.add(project)
    db.session.flush()

    db.session.commit()

    return jsonify({'success': True, 'data': project.to_dict()})


@app.route('/api/projects/<int:project_id>', methods=['PUT'])
def update_project(project_id):
    project = Project.query.get_or_404(project_id)
    data = request.json or {}
    name = (data.get('name') or '').strip()
    if name and name != project.name:
        if Project.query.filter_by(name=name).first():
            return jsonify({'success': False, 'message': '项目名称已存在'}), 400
        project.name = name
    project.description = data.get('description', project.description)
    db.session.commit()
    return jsonify({'success': True, 'data': project.to_dict()})


@app.route('/api/projects/<int:project_id>', methods=['DELETE'])
def delete_project(project_id):
    data = request.json or {}
    if data.get('password') != config.DELETE_PASSWORD:
        return jsonify({'success': False, 'message': '删除密码错误'}), 403

    project = Project.query.get_or_404(project_id)
    db.session.delete(project)
    db.session.commit()
    return jsonify({'success': True})


# ------------------- 版本 -------------------
@app.route('/api/projects/<int:project_id>/versions', methods=['GET'])
def get_versions(project_id):
    versions = Version.query.filter_by(project_id=project_id).order_by(Version.sort_order.asc(), Version.id.asc()).all()
    return jsonify({'success': True, 'data': [v.to_dict() for v in versions]})


@app.route('/api/projects/<int:project_id>/versions', methods=['POST'])
def create_version(project_id):
    data = request.json or {}
    name = (data.get('version_name') or '').strip()
    if not name:
        return jsonify({'success': False, 'message': '版本名称不能为空'}), 400
    if Version.query.filter_by(project_id=project_id, version_name=name).first():
        return jsonify({'success': False, 'message': '该项目下版本名已存在'}), 400
    max_order = db.session.query(db.func.max(Version.sort_order)).filter_by(project_id=project_id).scalar()
    version = Version(project_id=project_id, version_name=name, sort_order=(max_order if max_order is not None else -1) + 1)
    db.session.add(version)
    db.session.flush()
    init_system_columns(project_id, version.id)
    db.session.commit()
    return jsonify({'success': True, 'data': version.to_dict()})


@app.route('/api/projects/<int:project_id>/versions/<int:version_id>/copy', methods=['POST'])
def copy_version(project_id, version_id):
    source = Version.query.filter_by(id=version_id, project_id=project_id).first_or_404()
    init_system_columns(project_id, source.id)
    data = request.json or {}
    name = (data.get('version_name') or '').strip()
    if not name:
        return jsonify({'success': False, 'message': '副本名称不能为空'}), 400
    if Version.query.filter_by(project_id=project_id, version_name=name).first():
        return jsonify({'success': False, 'message': '该项目下版本名已存在'}), 400

    max_order = db.session.query(db.func.max(Version.sort_order)).filter_by(project_id=project_id).scalar()
    new_version = Version(
        project_id=project_id,
        version_name=name,
        sort_order=(max_order if max_order is not None else -1) + 1
    )
    try:
        db.session.add(new_version)
        db.session.flush()
        for source_column in query_version_columns(project_id, source.id):
            db.session.add(CustomColumn(
                project_id=project_id,
                version_id=new_version.id,
                name=source_column.name,
                key=source_column.key,
                is_system=source_column.is_system,
                is_visible=source_column.is_visible,
                width=source_column.width,
                sort_order=source_column.sort_order,
                text_align=source_column.text_align or 'left',
            ))
        source_cases = TestCase.query.filter_by(project_id=project_id, version_id=source.id) \
            .order_by(TestCase.sort_order.asc(), TestCase.id.asc()).all()
        case_id_map = {}
        image_replacements = []
        for source_case in source_cases:
            copied_case = TestCase(
                project_id=project_id,
                version_id=new_version.id,
                case_no=source_case.case_no,
                module=source_case.module,
                title=source_case.title,
                precondition=source_case.precondition,
                steps=source_case.steps,
                expected_result=source_case.expected_result,
                priority=source_case.priority,
                status=source_case.status,
                remark=source_case.remark,
                custom_fields=source_case.custom_fields,
                sort_order=source_case.sort_order,
            )
            db.session.add(copied_case)
            db.session.flush()
            case_id_map[source_case.id] = copied_case.id

            for source_image in source_case.images:
                image_data = source_image.image_data
                if not image_data:
                    continue
                copied_image = CaseImage(
                    test_case_id=copied_case.id,
                    filename=source_image.filename,
                    image_data=image_data,
                    mime_type=source_image.mime_type or 'application/octet-stream',
                )
                db.session.add(copied_image)
                image_replacements.append((copied_case, source_image, copied_image))

        db.session.flush()
        for copied_case, source_image, copied_image in image_replacements:
            new_src = f'/api/images/{copied_image.id}/content'
            tag_pattern = re.compile(r'<img\b[^>]*>', re.IGNORECASE)

            def replace_image_tag(match, old_id=source_image.id, new_id=copied_image.id, src=new_src):
                tag = match.group(0)
                has_id = re.search(r'data-image-id\s*=\s*["\']' + str(old_id) + r'["\']', tag, re.IGNORECASE)
                if not has_id:
                    return tag
                tag = re.sub(
                    r'(data-image-id\s*=\s*["\'])\d+(["\'])',
                    lambda m: m.group(1) + str(new_id) + m.group(2), tag, flags=re.IGNORECASE
                )
                tag = re.sub(
                    r'(src\s*=\s*["\'])[^"\']*(["\'])',
                    lambda m: m.group(1) + src + m.group(2), tag, flags=re.IGNORECASE
                )
                return tag

            copied_case.remark = tag_pattern.sub(replace_image_tag, copied_case.remark or '')

        for source_merge in CaseMerge.query.filter_by(project_id=project_id, version_id=source.id).all():
            copied_ids = [case_id_map[case_id] for case_id in source_merge.get_case_ids() if case_id in case_id_map]
            if len(copied_ids) >= 2:
                copied_merge = CaseMerge(
                    project_id=project_id,
                    version_id=new_version.id,
                    column_key=source_merge.column_key,
                )
                copied_merge.set_case_ids(copied_ids)
                db.session.add(copied_merge)

        db.session.commit()
        return jsonify({'success': True, 'data': {'version': new_version.to_dict(), 'copied_cases': len(source_cases)}})
    except Exception as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'创建版本副本失败: {str(exc)}'}), 500


@app.route('/api/projects/<int:project_id>/versions/order', methods=['POST'])
def update_version_order(project_id):
    data = request.json or {}
    orders = data.get('orders') or {}
    if not isinstance(orders, dict):
        return jsonify({'success': False, 'message': '版本排序数据格式错误'}), 400

    versions = Version.query.filter_by(project_id=project_id).all()
    version_map = {str(version.id): version for version in versions}
    try:
        requested = []
        for version_id, order in orders.items():
            version = version_map.get(str(version_id))
            if version is not None:
                requested.append((int(order), version.id, version))
        requested.sort(key=lambda item: (item[0], item[1]))
        ordered_ids = {item[1] for item in requested}
        remaining = sorted((version for version in versions if version.id not in ordered_ids),
                           key=lambda version: (version.sort_order or 0, version.id))
        for index, (_, _, version) in enumerate(requested):
            version.sort_order = index
        for index, version in enumerate(remaining, start=len(requested)):
            version.sort_order = index
        db.session.commit()
    except (TypeError, ValueError):
        db.session.rollback()
        return jsonify({'success': False, 'message': '版本排序值无效'}), 400
    return jsonify({'success': True, 'data': [v.to_dict() for v in sorted(versions, key=lambda v: (v.sort_order, v.id))]})


@app.route('/api/projects/<int:project_id>/versions/<int:version_id>', methods=['PUT'])
def update_version(project_id, version_id):
    version = Version.query.filter_by(id=version_id, project_id=project_id).first_or_404()
    data = request.json or {}
    name = (data.get('version_name') or '').strip()
    if not name:
        return jsonify({'success': False, 'message': '版本名称不能为空'}), 400
    if Version.query.filter(
        Version.project_id == project_id,
        Version.version_name == name,
        Version.id != version_id,
    ).first():
        return jsonify({'success': False, 'message': '该项目下版本名已存在'}), 400
    version.version_name = name
    db.session.commit()
    return jsonify({'success': True, 'data': version.to_dict()})


@app.route('/api/projects/<int:project_id>/versions/<int:version_id>', methods=['DELETE'])
def delete_version(project_id, version_id):
    data = request.json or {}
    if data.get('password') != config.DELETE_PASSWORD:
        return jsonify({'success': False, 'message': '删除密码错误'}), 403

    version = Version.query.filter_by(id=version_id, project_id=project_id).first_or_404()
    db.session.delete(version)
    db.session.commit()
    return jsonify({'success': True})


# ------------------- 自定义列 -------------------
@app.route('/api/projects/<int:project_id>/versions/<int:version_id>/columns', methods=['GET'])
def get_columns(project_id, version_id):
    Version.query.filter_by(id=version_id, project_id=project_id).first_or_404()
    init_system_columns(project_id, version_id)
    columns = query_version_columns(project_id, version_id)
    return jsonify({'success': True, 'data': [c.to_dict() for c in columns]})


@app.route('/api/projects/<int:project_id>/versions/<int:version_id>/columns', methods=['POST'])
def add_column(project_id, version_id):
    Version.query.filter_by(id=version_id, project_id=project_id).first_or_404()
    data = request.json or {}
    name = (data.get('name') or '').strip()
    key = (data.get('key') or '').strip()
    if not name or not key:
        return jsonify({'success': False, 'message': '列名和字段标识不能为空'}), 400
    if not key.isidentifier():
        return jsonify({'success': False, 'message': '字段标识需为合法标识符'}), 400
    if CustomColumn.query.filter_by(project_id=project_id, version_id=version_id, key=key).first():
        return jsonify({'success': False, 'message': '字段标识已存在'}), 400

    max_order = db.session.query(db.func.max(CustomColumn.sort_order)).filter_by(
        project_id=project_id, version_id=version_id
    ).scalar() or 0
    col = CustomColumn(
        project_id=project_id,
        version_id=version_id,
        name=name,
        key=key,
        is_system=False,
        is_visible=True,
        width=int(data.get('width', 150)),
        sort_order=max_order + 1
    )
    db.session.add(col)
    db.session.commit()
    return jsonify({'success': True, 'data': col.to_dict()})


@app.route('/api/columns/<int:column_id>', methods=['PUT'])
def update_column(column_id):
    col = CustomColumn.query.get_or_404(column_id)
    data = request.json or {}
    convert_to_system = (data.get('convert_to_system') or '').strip()
    system_def = next((item for item in SYSTEM_COLUMNS if item['key'] == convert_to_system), None)
    if convert_to_system:
        if col.is_system:
            return jsonify({'success': False, 'message': '系统列无需转换'}), 400
        if not system_def:
            return jsonify({'success': False, 'message': '目标系统列不存在'}), 400
        target = CustomColumn.query.filter_by(
            project_id=col.project_id, version_id=col.version_id,
            key=convert_to_system, is_system=True
        ).first()
        if not target:
            target = CustomColumn(
                project_id=col.project_id,
                version_id=col.version_id,
                name=system_def['name'],
                key=system_def['key'],
                is_system=True,
                is_visible=col.is_visible,
                width=system_def['width'],
                sort_order=col.sort_order,
                text_align=col.text_align or 'left',
            )
            db.session.add(target)
            db.session.flush()
        for case in TestCase.query.filter_by(
                project_id=col.project_id, version_id=col.version_id).all():
            custom = case.get_custom_fields()
            if col.key in custom:
                setattr(case, convert_to_system, custom[col.key])
                custom.pop(col.key, None)
                case.set_custom_fields(custom)
        db.session.delete(col)
        db.session.commit()
        return jsonify({'success': True, 'data': target.to_dict()})
    if 'name' in data:
        name = str(data.get('name') or '').strip()
        if not name:
            return jsonify({'success': False, 'message': '列名称不能为空'}), 400
        duplicate = CustomColumn.query.filter(
            CustomColumn.project_id == col.project_id,
            CustomColumn.version_id == col.version_id,
            CustomColumn.name == name,
            CustomColumn.id != col.id,
        ).first()
        if duplicate:
            return jsonify({'success': False, 'message': '当前版本已存在同名列'}), 400
        col.name = name
    if 'is_visible' in data:
        col.is_visible = bool(data['is_visible'])
    if 'width' in data:
        col.width = int(data['width'])
    if 'sort_order' in data:
        col.sort_order = int(data['sort_order'])
    if 'text_align' in data:
        text_align = (data.get('text_align') or '').strip().lower()
        if text_align not in {'left', 'center', 'right'}:
            return jsonify({'success': False, 'message': '对齐方式无效'}), 400
        col.text_align = text_align
    db.session.commit()
    return jsonify({'success': True, 'data': col.to_dict()})


@app.route('/api/columns/<int:column_id>', methods=['DELETE'])
def delete_column(column_id):
    col = CustomColumn.query.get_or_404(column_id)
    if col.is_system and col.key == 'status':
        return jsonify({'success': False, 'message': '执行结果列为固定列，不可删除'}), 400
    db.session.delete(col)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/projects/<int:project_id>/versions/<int:version_id>/columns/order', methods=['POST'])
def update_columns_order(project_id, version_id):
    Version.query.filter_by(id=version_id, project_id=project_id).first_or_404()
    data = request.json or {}
    orders = data.get('orders', {})
    for col_id, order in orders.items():
        col = CustomColumn.query.filter_by(
            id=col_id, project_id=project_id, version_id=version_id
        ).first()
        if col:
            col.sort_order = int(order)
    db.session.commit()
    return jsonify({'success': True})


# ------------------- 合并单元格 -------------------
@app.route('/api/projects/<int:project_id>/versions/<int:version_id>/merges', methods=['POST'])
def create_merge(project_id, version_id):
    data = request.json or {}
    column_key = (data.get('column_key') or '').strip()
    case_ids = data.get('case_ids') or []
    if column_key == 'case_no':
        return jsonify({'success': False, 'message': '用例编号按实际用例行自动编号，不参与合并'}), 400
    version = Version.query.filter_by(id=version_id, project_id=project_id).first_or_404()
    column = CustomColumn.query.filter_by(
        project_id=project_id, version_id=version_id, key=column_key
    ).first()
    if not column:
        return jsonify({'success': False, 'message': '列不存在'}), 400
    try:
        case_ids = list(dict.fromkeys(int(case_id) for case_id in case_ids))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': '合并范围无效'}), 400
    if len(case_ids) < 2:
        return jsonify({'success': False, 'message': '至少选择同一列的两个连续单元格'}), 400

    cases = TestCase.query.filter(
        TestCase.project_id == project_id,
        TestCase.version_id == version.id,
        TestCase.id.in_(case_ids)
    ).order_by(TestCase.sort_order.asc(), TestCase.id.asc()).all()
    if len(cases) != len(case_ids):
        return jsonify({'success': False, 'message': '合并单元格必须来自当前版本'}), 400
    all_cases = TestCase.query.filter_by(project_id=project_id, version_id=version_id) \
        .order_by(TestCase.sort_order.asc(), TestCase.id.asc()).all()
    positions = [all_cases.index(case) for case in cases]
    if positions != list(range(min(positions), max(positions) + 1)):
        return jsonify({'success': False, 'message': '只能合并同一列的连续单元格'}), 400

    for merge in CaseMerge.query.filter_by(project_id=project_id, version_id=version_id, column_key=column_key).all():
        if set(merge.get_case_ids()).intersection(case_ids):
            return jsonify({'success': False, 'message': '选中的单元格已经合并'}), 400

    merge = CaseMerge(project_id=project_id, version_id=version_id, column_key=column_key)
    merge.set_case_ids([case.id for case in cases])
    db.session.add(merge)
    db.session.flush()
    normalize_case_order(project_id, version_id, reset_numbers=True)
    db.session.commit()
    return jsonify({'success': True, 'data': merge.to_dict()})


@app.route('/api/merges/<int:merge_id>', methods=['DELETE'])
def delete_merge(merge_id):
    merge = CaseMerge.query.get_or_404(merge_id)
    project_id, version_id = merge.project_id, merge.version_id
    db.session.delete(merge)
    db.session.flush()
    normalize_case_order(project_id, version_id, reset_numbers=True)
    db.session.commit()
    return jsonify({'success': True})


def _case_display_sort_key(case):
    """兼容历史 sort_order 为空的数据，并保持列表当前可见顺序。"""
    if case.sort_order is None:
        return (1, case.id or 0, case.id or 0)
    return (0, case.sort_order, case.id or 0)


def logical_case_groups(project_id, version_id, cases=None):
    """按标题合并范围返回逻辑用例分组。

    Excel 中把一条用例的步骤拆成多行后，页面会通过合并“标题”把这些
    物理行标识为同一条逻辑用例。没有标题合并的版本仍按单行计算，避免
    影响正常格式导入的数据。
    """
    if cases is None:
        cases = TestCase.query.filter_by(
            project_id=project_id, version_id=version_id
        ).all()
    ordered_cases = sorted(cases, key=_case_display_sort_key)
    case_by_id = {case.id: case for case in ordered_cases}
    positions = {case.id: index for index, case in enumerate(ordered_cases)}
    merge_members = {}
    title_merges = CaseMerge.query.filter_by(
        project_id=project_id, version_id=version_id, column_key='title'
    ).all()
    for merge in title_merges:
        member_ids = [case_id for case_id in merge.get_case_ids()
                      if case_id in positions]
        if len(member_ids) < 2:
            continue
        member_ids.sort(key=positions.get)
        # 一个用例只能属于一个标题分组；异常重叠数据按先出现的合并保留。
        if any(case_id in merge_members for case_id in member_ids):
            continue
        for case_id in member_ids:
            merge_members[case_id] = member_ids

    groups = []
    seen = set()
    for case in ordered_cases:
        if case.id in seen:
            continue
        member_ids = merge_members.get(case.id, [case.id])
        group = [case_by_id[case_id] for case_id in member_ids
                 if case_id in case_by_id and case_id not in seen]
        if not group:
            continue
        groups.append(group)
        seen.update(item.id for item in group)
    return groups


def normalize_case_order(project_id, version_id, reset_numbers=False):
    """按当前列表顺序补齐行排序；插入后可同时重排用例编号。"""
    cases = TestCase.query.filter_by(project_id=project_id, version_id=version_id).all()
    cases.sort(key=_case_display_sort_key)
    for index, case in enumerate(cases, start=1):
        case.sort_order = index * 1000
    if reset_numbers:
        for index, group in enumerate(logical_case_groups(project_id, version_id, cases), start=1):
            for case in group:
                case.case_no = str(index)
    return cases


def ensure_case_numbers(project_id, version_id, cases=None):
    """按逻辑用例顺序校正编号；标题合并组只计为一条。"""
    if cases is None:
        cases = normalize_case_order(project_id, version_id)
    else:
        cases = sorted(cases, key=_case_display_sort_key)
    groups = logical_case_groups(project_id, version_id, cases)
    expected = [str(index) for index in range(1, len(groups) + 1)]
    current = [str(group[0].case_no or '') for group in groups]
    if current == expected:
        return False
    for group, case_no in zip(groups, expected):
        for case in group:
            case.case_no = case_no
    return True


def normalize_case_merges(project_id, version_id, cases=None):
    """修复插入行落在纵向合并区域中却未写入合并关系的历史数据。"""
    if cases is None:
        cases = TestCase.query.filter_by(project_id=project_id, version_id=version_id).all()
        cases.sort(key=_case_display_sort_key)
    positions = {case.id: index for index, case in enumerate(cases)}
    changed = False
    merges = CaseMerge.query.filter_by(project_id=project_id, version_id=version_id).all()
    for merge in merges:
        # 用例编号必须逐行显示并按真实用例数量连续编号。历史导入的
        # 编号合并只影响显示，不影响用例数据，直接清理其合并记录。
        if merge.column_key == 'case_no':
            db.session.delete(merge)
            changed = True
            continue
        member_ids = [case_id for case_id in merge.get_case_ids() if case_id in positions]
        if len(member_ids) < 2:
            continue
        start = min(positions[case_id] for case_id in member_ids)
        end = max(positions[case_id] for case_id in member_ids)
        expected_ids = [case.id for case in cases[start:end + 1]]
        if expected_ids != member_ids:
            merge.set_case_ids(expected_ids)
            changed = True
    return changed


def compute_sort_orders(project_id, version_id, target_id=None, position=None, count=1):
    """根据插入目标计算新行的 sort_order 列表（递增）。"""
    # 老数据可能没有 sort_order。先按原有列表顺序补齐，避免 None 参与算术运算，
    # 也让连续插入不会因为间隔耗尽而产生大量相同排序值。
    cases = normalize_case_order(project_id, version_id)
    n = max(1, count)
    if not cases:
        return [1000 * (i + 1) for i in range(n)]

    if not target_id or position not in ('above', 'below'):
        # 默认在顶部插入
        base = cases[0].sort_order
        return [base - (n - i) * 1000 for i in range(n)]

    idx = next((i for i, c in enumerate(cases) if c.id == target_id), -1)
    if idx < 0:
        base = cases[-1].sort_order
        return [base + (i + 1) * 1000 for i in range(n)]

    if position == 'above':
        prev_order = cases[idx - 1].sort_order if idx > 0 else None
        next_order = cases[idx].sort_order
        if prev_order is None:
            step = 1000
            start = next_order - (n + 1) * 1000
        else:
            gap = next_order - prev_order
            step = max(1, gap // (n + 1))
            start = prev_order
    else:  # below
        prev_order = cases[idx].sort_order
        next_order = cases[idx + 1].sort_order if idx + 1 < len(cases) else None
        if next_order is None:
            step = 1000
            start = prev_order
        else:
            gap = next_order - prev_order
            step = max(1, gap // (n + 1))
            start = prev_order

    return [start + step * (i + 1) for i in range(n)]


def normalize_merge_insert_target(project_id, version_id, target_id=None,
                                  position=None):
    """把合并区域内的插入点调整到整个合并块的边界。

    合并单元格不能被一条普通输入行从中间截断。用户在合并区域中的
    任意行选择“上方/下方插入”时，分别落到合并块首行上方或末行下方，
    这样新增后原有的 rowspan 和合并关系都能保持完整。
    """
    if target_id is None or position not in ('above', 'below'):
        return target_id, position

    cases = TestCase.query.filter_by(
        project_id=project_id, version_id=version_id
    ).all()
    cases.sort(key=_case_display_sort_key)
    positions = {case.id: index for index, case in enumerate(cases)}
    target_index = positions.get(target_id)
    if target_index is None:
        return target_id, position

    boundary_index = target_index
    merges = CaseMerge.query.filter_by(
        project_id=project_id, version_id=version_id
    ).all()
    for merge in merges:
        member_positions = [
            positions[case_id] for case_id in merge.get_case_ids()
            if case_id in positions
        ]
        if len(member_positions) < 2 or target_index not in member_positions:
            continue
        if position == 'above':
            boundary_index = min(boundary_index, min(member_positions))
        else:
            boundary_index = max(boundary_index, max(member_positions))

    return cases[boundary_index].id, position


# ------------------- 用例 -------------------
@app.route('/api/projects/<int:project_id>/versions/<int:version_id>/cases', methods=['GET'])
def get_cases(project_id, version_id):
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    if page_size not in [20, 50, 100]:
        page_size = 20
    keyword = (request.args.get('keyword', '') or '').strip()
    status_filter = (request.args.get('status', '') or '').strip()
    status_filters = [value for value in status_filter.split(',') if value in STATUS_LIST]

    init_system_columns(project_id, version_id)
    all_cases = TestCase.query.filter_by(project_id=project_id, version_id=version_id).all()
    if any(case.sort_order in (None, 0) for case in all_cases):
        all_cases = normalize_case_order(project_id, version_id)
    else:
        all_cases.sort(key=_case_display_sort_key)
    merges_changed = normalize_case_merges(project_id, version_id, all_cases)
    if merges_changed:
        db.session.flush()
    if merges_changed or ensure_case_numbers(project_id, version_id, all_cases):
        db.session.commit()
    query = TestCase.query.filter_by(project_id=project_id, version_id=version_id)
    if status_filters:
        query = query.filter(TestCase.status.in_(status_filters))
    if keyword:
        # 支持多个关键词，要求每个关键词都能在当前用例的任意字段中找到；
        # 同时覆盖系统字段和自定义字段，避免搜索结果过窄。
        searchable_fields = (
            TestCase.case_no,
            TestCase.module,
            TestCase.title,
            TestCase.precondition,
            TestCase.steps,
            TestCase.expected_result,
            TestCase.priority,
            TestCase.status,
            TestCase.remark,
            TestCase.custom_fields,
        )
        keywords = [part for part in re.split(r'\s+', keyword) if part]
        for part in keywords:
            query = query.filter(db.or_(*[
                field.contains(part, autoescape=True) for field in searchable_fields
            ]))

    total = query.count()
    cases = query.order_by(TestCase.sort_order.asc(), TestCase.id.asc()).offset((page - 1) * page_size).limit(page_size).all()
    columns = query_version_columns(project_id, version_id)
    columns_dict = [c.to_dict() for c in columns]
    merges = CaseMerge.query.filter_by(project_id=project_id, version_id=version_id).all()
    all_cases_by_id = {case.id: case for case in all_cases}
    system_column_keys = {column.key for column in columns if column.is_system}
    merge_data = []
    for merge in merges:
        item = merge.to_dict()
        # 合并单元格可能跨分页，前端仅凭当前页数据无法拼接完整内容。
        # 随合并关系返回各成员原始值，避免合并后只显示首行内容。
        values = {}
        for case_id in merge.get_case_ids():
            case = all_cases_by_id.get(case_id)
            if not case:
                continue
            if merge.column_key in system_column_keys:
                value = getattr(case, merge.column_key, '')
            else:
                value = case.get_custom_fields().get(merge.column_key, '')
            values[str(case_id)] = value or ''
        item['values'] = values
        merge_data.append(item)

    data = {
        'total': total,
        'page': page,
        'page_size': page_size,
        'columns': columns_dict,
        'cases': [c.to_dict(columns_dict) for c in cases],
        'merges': merge_data
    }
    return jsonify({'success': True, 'data': data})


def rich_value_to_excel_text(value):
    """把备注/富文本字段转换为 Excel 可读文本，并保留图片占位符。"""
    raw = str(value or '')
    if not raw:
        return ''
    if not re.search(r'<(?:img|br|div|p|s|strike|del)\b', raw, re.IGNORECASE):
        return html.unescape(raw)
    text_value = re.sub(r'<img\b[^>]*>', '[图片]', raw, flags=re.IGNORECASE)
    text_value = re.sub(r'<br\s*/?>', '\n', text_value, flags=re.IGNORECASE)
    text_value = re.sub(r'</(?:div|p|li)>', '\n', text_value, flags=re.IGNORECASE)
    text_value = re.sub(r'<[^>]+>', '', text_value)
    return html.unescape(text_value).strip()


def export_column_value(case, column):
    if column.is_system:
        return getattr(case, column.key, '')
    return (case.get_custom_fields() or {}).get(column.key, '')


def export_sheet_title(name):
    title = re.sub(r'[\\/*?:\[\]]', '_', str(name or '')).strip()[:31]
    return title or '用例列表'


def add_excel_image_in_cell(worksheet, image_data, row_index, column_index, image_index):
    """将图片以单元格锚点方式放入主表对应单元格。"""
    embedded = ExcelImage(io.BytesIO(bytes(image_data)))
    max_width, max_height = 96, 70
    scale = min(max_width / embedded.width, max_height / embedded.height, 1)
    width = max(1, int(embedded.width * scale))
    height = max(1, int(embedded.height * scale))
    column_offset = (image_index % 2) * 102 + 4
    row_offset = (image_index // 2) * 78 + 4
    embedded.width = width
    embedded.height = height
    embedded.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=column_index - 1,
            colOff=pixels_to_EMU(column_offset),
            row=row_index - 1,
            rowOff=pixels_to_EMU(row_offset),
        ),
        ext=XDRPositiveSize2D(cx=pixels_to_EMU(width), cy=pixels_to_EMU(height)),
    )
    worksheet.add_image(embedded)


@app.route('/api/projects/<int:project_id>/versions/<int:version_id>/export', methods=['GET'])
def export_version_excel(project_id, version_id):
    """导出当前版本的可见列、用例、合并关系和单元格内图片。"""
    version = Version.query.filter_by(id=version_id, project_id=project_id).first()
    project = Project.query.get(project_id)
    if not version or not project:
        return jsonify({'success': False, 'message': '项目或版本不存在'}), 404

    init_system_columns(project_id, version_id)
    columns = CustomColumn.query.filter_by(project_id=project_id, version_id=version_id, is_visible=True) \
        .order_by(CustomColumn.sort_order.asc(), CustomColumn.id.asc()).all()
    cases = TestCase.query.filter_by(project_id=project_id, version_id=version_id) \
        .order_by(TestCase.sort_order.asc(), TestCase.id.asc()).all()
    merges = CaseMerge.query.filter_by(project_id=project_id, version_id=version_id).all()

    workbook = Workbook()
    sheet_title = export_sheet_title(version.version_name)
    if sheet_title == '图片附件':
        sheet_title = '用例列表'
    worksheet = workbook.active
    worksheet.title = sheet_title
    worksheet.freeze_panes = 'A2'
    worksheet.sheet_view.showGridLines = False

    header_fill = PatternFill('solid', fgColor='409EFF')
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
    body_font = Font(name='Microsoft YaHei', size=10, color='1F2937')
    strike_font = Font(name='Microsoft YaHei', size=10, color='1F2937', strike=True)
    border = Border(
        left=Side(style='thin', color='D9E2F3'),
        right=Side(style='thin', color='D9E2F3'),
        top=Side(style='thin', color='D9E2F3'),
        bottom=Side(style='thin', color='D9E2F3'),
    )
    status_colors = {
        '通过': '67C23A',
        '失败': 'F56C6C',
        '未执行': '909399',
        '阻塞': 'E6A23C',
        '跳过': '409EFF',
    }

    for column_index, column in enumerate(columns, start=1):
        header = worksheet.cell(row=1, column=column_index, value=column.name)
        header.fill = header_fill
        header.font = header_font
        header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header.border = border
        # Web 端宽度是像素，Excel 列宽使用近似字符数。
        worksheet.column_dimensions[get_column_letter(column_index)].width = max(
            10, min(60, round((column.width or 120) / 7, 1))
        )
    worksheet.row_dimensions[1].height = 28

    case_row_map = {}
    column_index_map = {column.key: index for index, column in enumerate(columns, start=1)}
    image_column_index = column_index_map.get('remark') or (len(columns) if columns else None)
    for row_index, case in enumerate(cases, start=2):
        case_row_map[case.id] = row_index
        images = [image for image in case.images if image.image_data]
        has_image = bool(images)
        for column_index, column in enumerate(columns, start=1):
            raw_value = export_column_value(case, column)
            cell_value = rich_value_to_excel_text(raw_value)
            if column.key == 'remark' and has_image:
                # 图片直接锚定在备注单元格中，不再输出占位文字或附件工作表提示。
                cell_value = re.sub(r'\[图片\]', '', cell_value).strip()
            cell = worksheet.cell(row=row_index, column=column_index, value=cell_value)
            cell.border = border
            horizontal = column.text_align if column.text_align in {'left', 'center', 'right'} else 'left'
            cell.alignment = Alignment(horizontal=horizontal, vertical='top', wrap_text=True)
            cell.font = strike_font if re.search(r'<(?:s|strike|del)\b', str(raw_value or ''), re.IGNORECASE) else body_font
            if column.key == 'status':
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                status = normalize_status(raw_value)
                cell.font = Font(name='Microsoft YaHei', size=10, color=status_colors.get(status, '1F2937'))
        worksheet.row_dimensions[row_index].height = max(42, min(240, 78 * ((len(images) + 1) // 2)))
        if image_column_index:
            for image_index, image in enumerate(images):
                try:
                    add_excel_image_in_cell(worksheet, image.image_data, row_index, image_column_index, image_index)
                except Exception:
                    # 单张图片损坏时不影响其他用例导出。
                    continue

    if cases and columns:
        worksheet.auto_filter.ref = f'A1:{get_column_letter(len(columns))}{len(cases) + 1}'

    # 将当前版本的纵向合并关系还原到导出表格，隐藏列对应的合并自然跳过。
    for merge in merges:
        if merge.column_key == 'case_no':
            continue
        rows = [case_row_map[case_id] for case_id in merge.get_case_ids() if case_id in case_row_map]
        column_index = column_index_map.get(merge.column_key)
        if not column_index or len(rows) < 2:
            continue
        start_row, end_row = min(rows), max(rows)
        if sorted(rows) != list(range(start_row, end_row + 1)):
            continue
        worksheet.merge_cells(
            start_row=start_row,
            start_column=column_index,
            end_row=end_row,
            end_column=column_index,
        )
        merged_cell = worksheet.cell(row=start_row, column=column_index)
        merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    download_name = f'{project.name}_{version.version_name}_用例.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=download_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/api/projects/<int:project_id>/versions/<int:version_id>/cases/reset-numbers', methods=['POST'])
def reset_case_numbers(project_id, version_id):
    """按当前列表顺序将一个版本内的用例编号重排为 1..n。"""
    version = Version.query.filter_by(id=version_id, project_id=project_id).first()
    if not version:
        return jsonify({'success': False, 'message': '项目或版本不存在'}), 404

    cases = normalize_case_order(project_id, version_id, reset_numbers=True)
    db.session.commit()
    return jsonify({'success': True, 'data': {'reset': len(cases)}})


@app.route('/api/projects/<int:project_id>/versions/<int:version_id>/reset-status', methods=['POST'])
def reset_case_status(project_id, version_id):
    """将一个版本内所有用例的执行结果重置为未执行。"""
    version = Version.query.filter_by(id=version_id, project_id=project_id).first()
    if not version:
        return jsonify({'success': False, 'message': '项目或版本不存在'}), 404
    cases = TestCase.query.filter_by(project_id=project_id, version_id=version_id).all()
    for case in cases:
        case.status = '未执行'
    db.session.commit()
    return jsonify({'success': True, 'data': {'reset': len(cases)}})


@app.route('/api/cases', methods=['POST'])
def create_case():
    data = request.json or {}
    project_id = data.get('project_id')
    version_id = data.get('version_id')
    if not project_id or not version_id:
        return jsonify({'success': False, 'message': '项目或版本缺失'}), 400

    case_no = excel_value_to_text(data.get('case_no'))
    if not case_no:
        case_no = str(next_case_number(project_id, version_id))

    case = TestCase(
        project_id=project_id,
        version_id=version_id,
        case_no=case_no,
        module=data.get('module', ''),
        title=data.get('title', ''),
        precondition=data.get('precondition', ''),
        steps=data.get('steps', ''),
        expected_result=data.get('expected_result', ''),
        priority=data.get('priority', ''),
        status=data.get('status', '未执行'),
        remark=data.get('remark', ''),
        sort_order=compute_sort_orders(project_id, version_id, count=1)[0]
    )
    if case.status not in STATUS_LIST:
        case.status = '未执行'

    custom = data.get('custom_fields', {})
    if not isinstance(custom, dict):
        custom = {}
    valid_keys = {c.key for c in CustomColumn.query.filter_by(
        project_id=project_id, version_id=version_id, is_system=False).all()}
    case.set_custom_fields({k: v for k, v in (custom or {}).items() if k in valid_keys})

    db.session.add(case)
    ensure_case_numbers(project_id, version_id)
    db.session.commit()
    columns = query_version_columns(project_id, version_id)
    return jsonify({'success': True, 'data': case.to_dict([c.to_dict() for c in columns])})


@app.route('/api/cases/batch', methods=['POST'])
def create_cases_batch():
    data = request.json or {}
    project_id = data.get('project_id')
    version_id = data.get('version_id')
    cases_data = data.get('cases', [])
    if not project_id or not version_id:
        return jsonify({'success': False, 'message': '项目或版本缺失'}), 400
    if not cases_data:
        return jsonify({'success': False, 'message': '没有用例数据'}), 400

    target_id = data.get('insert_target')
    try:
        target_id = int(target_id) if target_id is not None else None
    except (TypeError, ValueError):
        target_id = None
    position = data.get('insert_position')  # 'above' 或 'below'
    valid_keys = {c.key for c in CustomColumn.query.filter_by(
        project_id=project_id, version_id=version_id, is_system=False).all()}
    target_id, position = normalize_merge_insert_target(
        project_id, version_id, target_id, position
    )
    sort_orders = compute_sort_orders(project_id, version_id, target_id, position, len(cases_data))
    next_number = next_case_number(project_id, version_id)

    created = []
    for i, item in enumerate(cases_data):
        item = item if isinstance(item, dict) else {}
        case_no = excel_value_to_text(item.get('case_no'))
        if not case_no:
            case_no = str(next_number)
            next_number += 1
        case = TestCase(
            project_id=project_id,
            version_id=version_id,
            case_no=case_no,
            module=item.get('module', ''),
            title=item.get('title', ''),
            precondition=item.get('precondition', ''),
            steps=item.get('steps', ''),
            expected_result=item.get('expected_result', ''),
            priority=item.get('priority', ''),
            status=item.get('status', '未执行'),
            remark=item.get('remark', ''),
            sort_order=sort_orders[i]
        )
        if case.status not in STATUS_LIST:
            case.status = '未执行'
        custom = item.get('custom_fields', {})
        if not isinstance(custom, dict):
            custom = {}
        case.set_custom_fields({k: v for k, v in (custom or {}).items() if k in valid_keys})
        db.session.add(case)
        created.append(case)

    db.session.flush()
    # 插入完成后立即按列表顺序重排排序值和编号，避免出现 1、2、98、43
    # 或历史空 sort_order 导致的新行位置错乱。
    ordered_cases = normalize_case_order(project_id, version_id, reset_numbers=True)
    inserted_ids = {case.id for case in created}
    if target_id and inserted_ids:
        # 只有插入块位于原合并区域的内部时，才把新行加入合并关系。
        # 插在合并首行上方或末行下方时，新行属于合并区域外，不能仅因
        # 目标行属于合并区域就扩大 rowspan，否则会造成页面列错位。
        positions = {case.id: index for index, case in enumerate(ordered_cases)}
        for merge in CaseMerge.query.filter_by(
                project_id=project_id, version_id=version_id).all():
            existing_ids = set(merge.get_case_ids())
            existing_positions = [positions[case_id] for case_id in existing_ids
                                  if case_id in positions]
            inserted_positions = [positions[case_id] for case_id in inserted_ids
                                  if case_id in positions]
            if (not existing_positions or not inserted_positions
                    or min(inserted_positions) <= min(existing_positions)
                    or max(inserted_positions) >= max(existing_positions)):
                continue
            merge.set_case_ids([
                case.id for case in ordered_cases
                if case.id in existing_ids or case.id in inserted_ids
            ])
    db.session.commit()
    columns = query_version_columns(project_id, version_id)
    columns_dict = [c.to_dict() for c in columns]
    return jsonify({'success': True, 'data': [c.to_dict(columns_dict) for c in created]})


@app.route('/api/cases/<int:case_id>', methods=['PUT'])
def update_case(case_id):
    case = TestCase.query.get_or_404(case_id)
    data = request.json or {}
    if 'case_no' in data:
        case_no = excel_value_to_text(data.get('case_no'))
        case.case_no = case_no or str(next_case_number(case.project_id, case.version_id))
    case.module = data.get('module', case.module)
    case.title = data.get('title', case.title)
    case.precondition = data.get('precondition', case.precondition)
    case.steps = data.get('steps', case.steps)
    case.expected_result = data.get('expected_result', case.expected_result)
    case.priority = data.get('priority', case.priority)
    if 'status' in data and data['status'] in STATUS_LIST:
        case.status = data['status']
    case.remark = data.get('remark', case.remark)

    custom = data.get('custom_fields', {}) or {}
    if not isinstance(custom, dict):
        custom = {}
    with db.session.no_autoflush:
        valid_keys = {c.key for c in CustomColumn.query.filter_by(
            project_id=case.project_id, version_id=case.version_id, is_system=False).all()}
    merged = case.get_custom_fields()
    merged.update({k: v for k, v in custom.items() if k in valid_keys})
    case.set_custom_fields(merged)

    db.session.commit()
    columns = query_version_columns(case.project_id, case.version_id)
    return jsonify({'success': True, 'data': case.to_dict([c.to_dict() for c in columns])})


@app.route('/api/cases/<int:case_id>', methods=['DELETE'])
def delete_case(case_id):
    case = TestCase.query.get_or_404(case_id)
    for merge in CaseMerge.query.filter_by(project_id=case.project_id, version_id=case.version_id).all():
        remaining_ids = [value for value in merge.get_case_ids() if value != case.id]
        if len(remaining_ids) < 2:
            db.session.delete(merge)
        else:
            merge.set_case_ids(remaining_ids)
    db.session.delete(case)
    db.session.flush()
    normalize_case_order(case.project_id, case.version_id, reset_numbers=True)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/cases/batch-delete', methods=['POST'])
def delete_cases_batch():
    data = request.json or {}
    raw_ids = data.get('case_ids') or []
    if not isinstance(raw_ids, list):
        return jsonify({'success': False, 'message': '用例编号格式错误'}), 400
    try:
        case_ids = list(dict.fromkeys(int(case_id) for case_id in raw_ids))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': '用例编号格式错误'}), 400
    cases = TestCase.query.filter(TestCase.id.in_(case_ids)).all() if case_ids else []
    for case in cases:
        for merge in CaseMerge.query.filter_by(project_id=case.project_id, version_id=case.version_id).all():
            remaining_ids = [value for value in merge.get_case_ids() if value != case.id]
            if len(remaining_ids) < 2:
                db.session.delete(merge)
            else:
                merge.set_case_ids(remaining_ids)
        db.session.delete(case)
    db.session.flush()
    for project_id, version_id in {
        (case.project_id, case.version_id) for case in cases
    }:
        normalize_case_order(project_id, version_id, reset_numbers=True)
    db.session.commit()
    return jsonify({'success': True, 'data': {'deleted': len(cases)}})


# ------------------- 统计 -------------------
@app.route('/api/projects/<int:project_id>/versions/<int:version_id>/stats', methods=['GET'])
def get_stats(project_id, version_id):
    all_cases = TestCase.query.filter_by(project_id=project_id, version_id=version_id).all()
    logical_cases = [group[0] for group in logical_case_groups(
        project_id, version_id, all_cases
    ) if group]
    total = len(logical_cases)
    stats = {}
    for status in STATUS_LIST:
        count = sum(1 for case in logical_cases if case.status == status)
        stats[status] = {
            'count': count,
            'percent': round(count / total * 100, 1) if total else 0
        }
    return jsonify({
        'success': True,
        'data': {
            'total': total,
            'stats': stats
        }
    })


@app.route('/api/projects/<int:project_id>/versions/<int:version_id>/summary', methods=['GET'])
def get_summary(project_id, version_id):
    all_cases = TestCase.query.filter_by(project_id=project_id, version_id=version_id).all()
    logical_cases = [group[0] for group in logical_case_groups(
        project_id, version_id, all_cases
    ) if group]
    total = len(logical_cases)
    counts = {}
    for status in STATUS_LIST:
        counts[status] = sum(1 for case in logical_cases if case.status == status)

    executed = total - counts['跳过'] - counts['阻塞']
    fail_cases = [case for case in logical_cases if case.status == '失败']
    skip_cases = [case for case in logical_cases if case.status == '跳过']
    block_cases = [case for case in logical_cases if case.status == '阻塞']

    def summary_text(value):
        """把富文本备注转换为总结用纯文本，图片只从总结中排除。"""
        value = str(value or '')
        value = re.sub(r'<img\b[^>]*>', '', value, flags=re.IGNORECASE)
        value = re.sub(r'<br\s*/?>', '\n', value, flags=re.IGNORECASE)
        value = re.sub(r'</(?:div|p)>', '\n', value, flags=re.IGNORECASE)
        value = re.sub(r'<[^>]+>', '', value)
        value = html.unescape(value)
        value = re.sub(r'[ \t]+\n', '\n', value)
        return value.strip()

    def reasons(cases):
        result = []
        for c in cases:
            remark = summary_text(c.remark)
            result.append({
                'id': c.id,
                'case_no': c.case_no,
                'title': c.title,
                'reason': remark or '未填写问题描述',
            })
        return result

    return jsonify({
        'success': True,
        'data': {
            'total': total,
            'executed': executed,
            'success': counts['通过'],
            'fail': counts['失败'],
            'block': counts['阻塞'],
            'skip': counts['跳过'],
            'unexecuted': counts['未执行'],
            'fail_reasons': reasons(fail_cases),
            'skip_reasons': reasons(skip_cases),
            'block_reasons': reasons(block_cases),
            'can_summarize': counts['未执行'] == 0 and total > 0
        }
    })


# ------------------- Excel 导入 -------------------
def _xml_local_name(tag):
    return tag.rsplit('}', 1)[-1]


def _xml_attr(element, name):
    for key, value in element.attrib.items():
        if _xml_local_name(key) == name:
            return value
    return None


def _active_worksheet_xml_path(archive):
    """返回工作簿当前活动工作表的 XML 路径。"""
    workbook = ET.fromstring(archive.read('xl/workbook.xml'))
    sheets = [node for node in workbook.iter() if _xml_local_name(node.tag) == 'sheet']
    active_tab = 0
    for node in workbook.iter():
        if _xml_local_name(node.tag) == 'workbookView':
            try:
                active_tab = int(node.attrib.get('activeTab', 0))
            except (TypeError, ValueError):
                active_tab = 0
            break
    sheet = sheets[min(max(active_tab, 0), max(len(sheets) - 1, 0))]
    relation_id = _xml_attr(sheet, 'id')
    rels = ET.fromstring(archive.read('xl/_rels/workbook.xml.rels'))
    target = None
    for relation in rels.iter():
        if relation.attrib.get('Id') == relation_id:
            target = relation.attrib.get('Target')
            break
    if not target:
        raise ValueError('无法定位 Excel 活动工作表')
    if target.startswith('/'):
        return target.lstrip('/')
    return posixpath.normpath(posixpath.join('xl', target))


def inspect_excel_bounds(file_storage):
    """快速扫描工作表 XML，只找真实有值的行和合并范围。

    某些 Excel 文件会把整张表写成带格式的空行/空列，openpyxl 的
    max_row/max_column 会因此膨胀。这里不读取单元格值，只扫描 XML 的
    有值节点，避免后续 iter_rows 遍历几十万空单元格。
    """
    stream = getattr(file_storage, 'stream', file_storage)
    stream.seek(0)
    try:
        with ZipFile(stream) as archive:
            sheet_path = _active_worksheet_xml_path(archive)
            last_value_row = 1
            merge_refs = []
            # 只按 XML 行块扫描，不让 openpyxl/ElementTree 为数万条格式空行
            # 创建对象。带值单元格一定包含 v、is 或 f 节点；只有样式的空行会被跳过。
            worksheet_bytes = archive.read(sheet_path)
            row_start = 0
            while True:
                row_start = worksheet_bytes.find(b'<row ', row_start)
                if row_start < 0:
                    break
                row_end = worksheet_bytes.find(b'</row>', row_start)
                if row_end < 0:
                    break
                row_number_start = worksheet_bytes.find(b' r="', row_start, row_end)
                if row_number_start >= 0:
                    row_number_start += 4
                    row_number_end = worksheet_bytes.find(b'"', row_number_start, row_end)
                    row_number = int(worksheet_bytes[row_number_start:row_number_end])
                    row_body = worksheet_bytes[row_start:row_end]
                    if (b'<v' in row_body or b'<is' in row_body or b'<f' in row_body):
                        last_value_row = max(last_value_row, row_number)
                row_start = row_end + 6

            merge_start = 0
            while True:
                merge_start = worksheet_bytes.find(b'<mergeCell ', merge_start)
                if merge_start < 0:
                    break
                ref_start = worksheet_bytes.find(b' ref="', merge_start)
                if ref_start >= 0:
                    ref_start += 6
                    ref_end = worksheet_bytes.find(b'"', ref_start)
                    if ref_end > ref_start:
                        merge_refs.append(worksheet_bytes[ref_start:ref_end].decode('utf-8'))
                merge_start += 11

            merge_ranges = []
            for reference in merge_refs:
                try:
                    merge_range = CellRange(reference)
                except ValueError:
                    continue
                merge_ranges.append(merge_range)
                last_value_row = max(last_value_row, merge_range.max_row)
            return last_value_row, merge_ranges
    finally:
        stream.seek(0)


def safe_custom_key(name):
    """生成不会与系统字段冲突的自定义列 key"""
    # 统一使用稳定的 ASCII key，避免表头包含特殊字符、过长或重复时
    # 无法作为前端字段标识的问题。
    normalized = re.sub(r'[^0-9A-Za-z_]+', '_', name).strip('_').lower()
    if not normalized:
        normalized = 'column'
    if normalized[0].isdigit():
        normalized = f'column_{normalized}'
    return f"c_{normalized}"


IMPORT_SYSTEM_HEADER_ALIASES = {
    '用例编号': 'case_no',
    '用例序号': 'case_no',
    '编号': 'case_no',
    'case_no': 'case_no',
    'case no': 'case_no',
    '模块': 'module',
    'module': 'module',
    '标题': 'title',
    '用例标题': 'title',
    'title': 'title',
    '前置条件': 'precondition',
    'precondition': 'precondition',
    '步骤': 'steps',
    '操作步骤': 'steps',
    '测试步骤': 'steps',
    'steps': 'steps',
    '预期结果': 'expected_result',
    '期望结果': 'expected_result',
    'expected_result': 'expected_result',
    '优先级': 'priority',
    'priority': 'priority',
    '执行结果': 'status',
    '测试结果': 'status',
    '状态': 'status',
    'status': 'status',
    '备注': 'remark',
    '说明': 'remark',
    'remark': 'remark',
}


def excel_value_to_text(value):
    """将 Excel 单元格值转换为适合保存到文本字段的内容。"""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def excel_cell_to_text(cell, preserve_strike=True):
    """读取单元格内容，并把 Excel 的删除线样式保存为安全的 HTML 标记。"""
    value = cell.value if hasattr(cell, 'value') else cell
    text_value = excel_value_to_text(value)
    if not text_value or not preserve_strike:
        return text_value
    font = getattr(cell, 'font', None)
    if font is not None and font.strike:
        return f'<s>{html.escape(text_value, quote=False)}</s>'
    return text_value


class LegacyExcelCell:
    """给旧版 .xls 单元格提供 openpyxl 读取代码所需的最小接口。"""

    def __init__(self, value, strike=False):
        self.value = value
        self.font = type('LegacyFont', (), {'strike': strike})()


def load_legacy_xls(raw_bytes):
    """读取旧版二进制 .xls，转换为当前导入流程使用的单元格结构。"""
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError('检测到旧版 .xls 文件，请先执行 pip install -r requirements.txt') from exc

    try:
        book = xlrd.open_workbook(file_contents=raw_bytes, formatting_info=True)
        sheet = book.sheet_by_index(0)
    except Exception as exc:
        raise ValueError('Excel 文件损坏，或文件扩展名与实际格式不一致') from exc

    def cell_at(row_index, col_index):
        cell = sheet.cell(row_index, col_index)
        strike = False
        try:
            xf = book.xf_list[cell.xf_index]
            strike = bool(book.font_list[xf.font_index].struck_out)
        except Exception:
            pass
        return LegacyExcelCell(cell.value, strike)

    rows = [
        [cell_at(row_index, col_index) for col_index in range(sheet.ncols)]
        for row_index in range(sheet.nrows)
    ]
    last_data_row = max(
        (row_index + 1 for row_index, row in enumerate(rows)
         if any(cell.value is not None and str(cell.value).strip() for cell in row)),
        default=1
    )
    merge_ranges = []
    for row_start, row_end, col_start, col_end in sheet.merged_cells:
        merge_ranges.append(CellRange(
            min_col=col_start + 1,
            min_row=row_start + 1,
            max_col=col_end,
            max_row=row_end,
        ))
        last_data_row = max(last_data_row, row_end)
    return book, rows, last_data_row, merge_ranges


def next_case_number(project_id, version_id):
    """返回该版本下下一个可用的数字用例编号。"""
    case_nos = db.session.query(TestCase.case_no).filter_by(
        project_id=project_id, version_id=version_id
    ).all()
    numbers = []
    for (case_no,) in case_nos:
        value = str(case_no or '').strip()
        if value.isdigit():
            numbers.append(int(value))
    return max(numbers, default=0) + 1


STATUS_ALIASES = {
    'pass': '通过', 'passed': '通过', 'success': '通过', '成功': '通过',
    'fail': '失败', 'failed': '失败', 'failure': '失败', '失败': '失败',
    'not run': '未执行', 'notrun': '未执行', 'pending': '未执行', '未执行': '未执行',
    'blocked': '阻塞', 'block': '阻塞', '阻塞': '阻塞',
    'skip': '跳过', 'skipped': '跳过', '跳过': '跳过',
    '通过': '通过',
}


def normalize_status(value):
    """统一 Excel/历史数据中常见的英文和中文执行结果。"""
    text_value = excel_value_to_text(value)
    return STATUS_ALIASES.get(text_value.casefold(), text_value if text_value in STATUS_LIST else '未执行')


@app.route('/api/projects/<int:project_id>/versions/<int:version_id>/import', methods=['POST'])
def import_cases(project_id, version_id):
    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'message': '未上传文件'}), 400

    wb = None
    legacy_book = None
    try:
        version = Version.query.filter_by(id=version_id, project_id=project_id).first()
        if not version:
            return jsonify({'success': False, 'message': '项目或版本不存在'}), 404
        init_system_columns(project_id, version_id)

        # Werkzeug 的部分上传流（尤其是 SpooledTemporaryFile 包装对象）不一定
        # 实现 seekable()，先复制到标准 BytesIO。根据文件头区分 .xlsx 和旧版 .xls。
        raw_bytes = file.read()
        if raw_bytes.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'):
            legacy_book, legacy_rows, last_data_row, worksheet_merges = load_legacy_xls(raw_bytes)
            header_row = legacy_rows[0] if legacy_rows else ()
            rows = legacy_rows[:last_data_row]
        else:
            if not raw_bytes.startswith(b'PK'):
                raise ValueError('上传文件不是有效的 Excel 文件，请选择 .xlsx 或 .xls 文件')
            excel_stream = io.BytesIO(raw_bytes)
            try:
                last_data_row, worksheet_merges = inspect_excel_bounds(excel_stream)
                wb = openpyxl.load_workbook(excel_stream, read_only=True, data_only=False)
            except (BadZipFile, KeyError, ValueError, OSError) as exc:
                raise ValueError(
                    'Excel 文件格式不完整或已损坏，请重新保存为标准 .xlsx 后再导入'
                ) from exc
            ws = wb.active
            header_row = next(
                ws.iter_rows(min_row=1, max_row=1, max_col=ws.max_column, values_only=True),
                ()
            )
            rows = [header_row]
        if not header_row:
            return jsonify({'success': False, 'message': 'Excel 为空或缺少表头'}), 400

        # 第一行是唯一的列定义。截掉第一行末尾的空单元格，避免 Excel
        # 的格式残留被误识别为额外列；表头中间出现空列则直接提示用户。
        raw_headers = list(header_row)
        last_header_idx = max(
            (idx for idx, value in enumerate(raw_headers) if value is not None and str(value).strip()),
            default=-1
        )
        if last_header_idx < 0:
            return jsonify({'success': False, 'message': 'Excel 第一行缺少表头'}), 400
        headers = [excel_value_to_text(value) for value in raw_headers[:last_header_idx + 1]]
        empty_headers = [str(idx + 1) for idx, header in enumerate(headers) if not header]
        if empty_headers:
            return jsonify({'success': False, 'message': f'第 {"、".join(empty_headers)} 列表头不能为空'}), 400
        if len(set(headers)) != len(headers):
            return jsonify({'success': False, 'message': '第一行存在重复表头，请先修改后再导入'}), 400

        # .xlsx 已按表头列数读取；.xls 则在这里截掉格式残留列。
        rows = [list(row[:len(headers)]) for row in rows]
        if wb is not None and last_data_row >= 2:
            rows.extend(ws.iter_rows(
                min_row=2,
                max_row=last_data_row,
                max_col=len(headers),
                values_only=False
            ))

        # 系统字段按表头自动映射，其余表头自动创建为自定义列。
        mapped_headers = [(idx, header, IMPORT_SYSTEM_HEADER_ALIASES.get(header.lower()))
                          for idx, header in enumerate(headers)]
        system_indexes = {mapped for _, _, mapped in mapped_headers if mapped}
        custom_cols = [(idx, header) for idx, header, mapped in mapped_headers if not mapped]

        existing_custom = {c.key: c for c in CustomColumn.query.filter_by(
            project_id=project_id, version_id=version_id, is_system=False).all()}
        existing_custom_by_name = {c.name: c for c in existing_custom.values()}
        max_order = db.session.query(db.func.max(CustomColumn.sort_order)).filter_by(
            project_id=project_id, version_id=version_id).scalar() or 0
        imported_keys = []
        imported_custom_keys = {}
        for idx, h in custom_cols:
            existing_named = existing_custom_by_name.get(h)
            key = existing_named.key if existing_named else safe_custom_key(h)
            # 兼容历史数据中已经存在的同名列，同时避免不同表头转换成同一个 key。
            original_key = key
            suffix = 2
            while key in existing_custom and existing_custom[key].name != h:
                key = f'{original_key}_{suffix}'
                suffix += 1
            imported_keys.append(key)
            imported_custom_keys[h] = key
            if key not in existing_custom:
                col = CustomColumn(
                    project_id=project_id,
                    version_id=version_id,
                    name=h,
                    key=key,
                    is_system=False,
                    is_visible=True,
                    width=150,
                    sort_order=max_order + 1
                )
                db.session.add(col)
                existing_custom[key] = col
                existing_custom_by_name[h] = col
                max_order += 1
        db.session.flush()

        # 导入后让表格列与 Excel 表头一致；执行结果仍保留为固定列。
        all_cols = CustomColumn.query.filter_by(
            project_id=project_id, version_id=version_id).all()
        system_cols = [c for c in all_cols if c.is_system]
        for c in system_cols:
            c.is_visible = c.key in system_indexes or c.key == 'status'
        for c in existing_custom.values():
            c.is_visible = c.key in imported_keys

        # 导入列的排序也按 Excel 表头排列；未出现在表头中的固定执行结果列放在最后。
        columns_by_key = {c.key: c for c in all_cols}
        imported_column_keys = []
        for _, header, system_key in mapped_headers:
            key = system_key or imported_custom_keys.get(header)
            if key and key in columns_by_key:
                if system_key:
                    columns_by_key[key].name = header
                columns_by_key[key].sort_order = len(imported_column_keys)
                imported_column_keys.append(key)
        next_order = len(imported_column_keys)
        for c in sorted(all_cols, key=lambda col: (col.key != 'status', col.sort_order, col.id)):
            if c.key not in imported_column_keys:
                c.sort_order = next_order
                next_order += 1

        # 导入数据
        created_count = 0
        next_number = next_case_number(project_id, version_id)
        # 保存 Excel 行号，后续用它把工作表中的合并范围映射回导入后的用例。
        vertical_merges = [merged for merged in worksheet_merges
                           if merged.min_row >= 2
                           and merged.min_col == merged.max_col
                           and merged.max_col <= len(headers)]
        merged_row_numbers = {
            row_number
            for merged in vertical_merges
            for row_number in range(merged.min_row, merged.max_row + 1)
        }

        data_rows = []
        for excel_row_number, row in enumerate(rows[1:], start=2):
            values = list(row[:len(headers)])
            # 合并单元格的非首行通常没有值；即使整行为空，也要保留为一条用例，
            # 否则无法还原 Excel 中的纵向合并关系。
            if (not any(cell.value is not None and str(cell.value).strip() for cell in values)
                    and excel_row_number not in merged_row_numbers):
                continue
            data_rows.append((excel_row_number, values))

        sort_orders = compute_sort_orders(project_id, version_id, count=len(data_rows)) if data_rows else []
        imported_case_ids_by_row = {}
        for row_idx, (excel_row_number, row) in enumerate(data_rows):
            custom_fields = {}
            for idx, h in custom_cols:
                key = imported_custom_keys[h]
                custom_fields[key] = excel_cell_to_text(row[idx]) if idx < len(row) else ''

            values_by_system_key = {}
            for idx, header, system_key in mapped_headers:
                if system_key:
                    # 执行结果需要保持为纯文本供状态选择器识别，其余字段保留删除线。
                    values_by_system_key[system_key] = excel_cell_to_text(
                        row[idx], preserve_strike=system_key not in {'status'}
                    ) if idx < len(row) else ''

            status = normalize_status(values_by_system_key.get('status', '未执行'))

            case = TestCase(
                project_id=project_id,
                version_id=version_id,
                # 表格未填写编号时按导入顺序自动生成 1、2、3……；已有数据则从最大数字继续。
                case_no=values_by_system_key.get('case_no') or str(next_number),
                module=values_by_system_key.get('module', ''),
                title=values_by_system_key.get('title', ''),
                precondition=values_by_system_key.get('precondition', ''),
                steps=values_by_system_key.get('steps', ''),
                expected_result=values_by_system_key.get('expected_result', ''),
                priority=values_by_system_key.get('priority', ''),
                status=status,
                remark=values_by_system_key.get('remark', ''),
                sort_order=sort_orders[row_idx]
            )
            case.set_custom_fields(custom_fields)
            db.session.add(case)
            db.session.flush()
            imported_case_ids_by_row[excel_row_number] = case.id
            created_count += 1
            if not values_by_system_key.get('case_no'):
                next_number += 1

        # 自动还原 Excel 中同一列的纵向合并。横向合并无法映射到当前用例表的
        # “同一列跨行”模型，因此保留原数据但不创建错误的合并记录。
        imported_keys_by_column = {
            index + 1: (system_key or imported_custom_keys.get(header))
            for index, header, system_key in mapped_headers
        }
        for merged in vertical_merges:
            column_key = imported_keys_by_column.get(merged.min_col)
            # 用例编号按真实用例行编号，不能继承 Excel 中的纵向编号合并。
            if column_key == 'case_no':
                continue
            case_ids = [
                imported_case_ids_by_row[row_number]
                for row_number in range(merged.min_row, merged.max_row + 1)
                if row_number in imported_case_ids_by_row
            ]
            if not column_key or len(case_ids) < 2:
                continue
            already_merged = False
            for existing_merge in CaseMerge.query.filter_by(
                    project_id=project_id, version_id=version_id, column_key=column_key).all():
                if set(existing_merge.get_case_ids()).intersection(case_ids):
                    already_merged = True
                    break
            if not already_merged:
                imported_merge = CaseMerge(
                    project_id=project_id,
                    version_id=version_id,
                    column_key=column_key
                )
                imported_merge.set_case_ids(case_ids)
                db.session.add(imported_merge)

        # 导入后按真实用例行重置编号，Excel 中的合并行不参与计数。
        normalize_case_order(project_id, version_id, reset_numbers=True)
        db.session.commit()
        return jsonify({'success': True, 'data': {'imported': created_count}})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'}), 500
    finally:
        if wb is not None:
            wb.close()
        if legacy_book is not None:
            legacy_book.release_resources()


# ------------------- 图片上传 -------------------
@app.route('/api/cases/<int:case_id>/images', methods=['POST'])
def upload_image(case_id):
    case = TestCase.query.get_or_404(case_id)
    files = request.files.getlist('images')
    saved = []
    for file in files:
        if not file or not file.filename:
            continue
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp']:
            continue
        image_data = file.read()
        if not image_data:
            continue
        img = CaseImage(
            test_case_id=case_id,
            filename=file.filename,
            image_data=image_data,
            mime_type=file.mimetype or 'application/octet-stream',
        )
        db.session.add(img)
        saved.append(img)
    db.session.commit()
    return jsonify({'success': True, 'data': [i.to_dict() for i in saved]})


@app.route('/api/cases/<int:case_id>/images', methods=['GET'])
def get_images(case_id):
    images = CaseImage.query.filter_by(test_case_id=case_id).all()
    return jsonify({'success': True, 'data': [i.to_dict() for i in images]})


@app.route('/api/images/<int:image_id>', methods=['DELETE'])
def delete_image(image_id):
    img = CaseImage.query.get_or_404(image_id)
    db.session.delete(img)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/images/<int:image_id>/content', methods=['GET'])
def serve_image_content(image_id):
    """从 MySQL 读取图片本体。"""
    img = CaseImage.query.get_or_404(image_id)
    if img.image_data:
        return send_file(
            io.BytesIO(bytes(img.image_data)),
            mimetype=img.mime_type or 'application/octet-stream',
            download_name=img.filename or 'image',
            max_age=31536000,
        )
    return jsonify({'success': False, 'message': '图片内容不存在'}), 404


# ------------------- 数据库备份 -------------------
@app.route('/api/backup', methods=['POST'])
def backup_database():
    data = request.json or {}
    backup_dir = data.get('backup_dir', config.BACKUP_DIR)
    try:
        os.makedirs(backup_dir, exist_ok=True)
    except Exception as e:
        return jsonify({'success': False, 'message': f'无法创建备份目录: {str(e)}'}), 400

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"case_manager_backup_{timestamp}.sql"
    filepath = os.path.join(backup_dir, filename)

    # 优先使用 mysqldump
    mysqldump = shutil.which('mysqldump')
    if mysqldump:
        cmd = [
            mysqldump,
            '-h', config.DB_HOST,
            '-P', str(config.DB_PORT),
            '-u', config.DB_USER,
            f'--password={config.DB_PASSWORD}',
            '--single-transaction',
            '--routines',
            '--triggers',
            config.DB_NAME
        ]
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                subprocess.run(cmd, stdout=f, check=True)
            return jsonify({'success': True, 'data': {'path': filepath}})
        except subprocess.CalledProcessError as e:
            return jsonify({'success': False, 'message': f'mysqldump 失败: {str(e)}'}), 500
    else:
        # 降级：导出核心表为 JSON
        try:
            import_data = {
                'projects': [p.to_dict() for p in Project.query.all()],
                'versions': [v.to_dict() for v in Version.query.all()],
                'columns': [c.to_dict() for c in CustomColumn.query.all()],
                'cases': [c.to_dict() for c in TestCase.query.all()],
                'merges': [m.to_dict() for m in CaseMerge.query.all()],
            }
            with open(filepath.replace('.sql', '.json'), 'w', encoding='utf-8') as f:
                json.dump(import_data, f, ensure_ascii=False, indent=2)
            return jsonify({'success': True, 'data': {'path': filepath.replace('.sql', '.json'), 'note': '未找到 mysqldump，已导出为 JSON'}})
        except Exception as e:
            return jsonify({'success': False, 'message': f'备份失败: {str(e)}'}), 500


# ------------------- 初始化 -------------------
@app.cli.command('init-db')
def init_db_command():
    create_database()
    initialize_auth_data()
    migrate_sort_order()
    print('数据库初始化完成')


def migrate_sort_order():
    """为已有数据补充排序字段，并将图片存储字段升级到数据库。"""
    try:
        with db.engine.connect() as conn:
            for column, definition in (
                ('precondition', "TEXT"),
                ('expected_result', "TEXT"),
                ('priority', "VARCHAR(50) DEFAULT ''"),
            ):
                try:
                    conn.execute(text(f"ALTER TABLE test_cases ADD COLUMN {column} {definition}"))
                except Exception:
                    pass  # 字段已存在
            try:
                conn.execute(text("ALTER TABLE custom_columns ADD COLUMN text_align VARCHAR(10) NOT NULL DEFAULT 'left'"))
            except Exception:
                pass  # 字段已存在
            conn.execute(text("UPDATE custom_columns SET text_align = 'left' WHERE text_align IS NULL OR text_align = ''"))
            try:
                conn.execute(text("ALTER TABLE test_cases ADD COLUMN sort_order INT DEFAULT 0"))
            except Exception:
                pass  # 字段已存在
            conn.execute(text("UPDATE test_cases SET sort_order = id * 1000 WHERE sort_order IS NULL OR sort_order = 0"))
            version_sort_added = False
            try:
                conn.execute(text("ALTER TABLE versions ADD COLUMN sort_order INT DEFAULT 0"))
                version_sort_added = True
            except Exception:
                pass  # 字段已存在
            # 新上传图片保存为 MySQL LONGBLOB。
            for column, definition in (
                ('image_data', "LONGBLOB NULL"),
                ('mime_type', "VARCHAR(100) NOT NULL DEFAULT 'application/octet-stream'"),
            ):
                try:
                    conn.execute(text(f"ALTER TABLE case_images ADD COLUMN {column} {definition}"))
                except Exception:
                    pass  # 字段已存在
            # 首次迁移时保持原来按创建时间倒序的显示顺序；后续启动不触碰用户排序。
            if version_sort_added:
                rows = conn.execute(text(
                    "SELECT id FROM versions ORDER BY created_at DESC, id DESC"
                )).fetchall()
                for offset, row in enumerate(rows):
                    conn.execute(text("UPDATE versions SET sort_order = :sort_order WHERE id = :id"),
                                 {'sort_order': offset, 'id': row[0]})
            conn.commit()
    except Exception as e:
        print('sort_order 迁移提示:', e)
    migrate_version_columns()


def migrate_version_columns():
    """把历史项目级列迁移为版本级列，并修复导入造成的跨版本覆盖。"""
    try:
        with db.engine.begin() as conn:
            try:
                conn.execute(text(
                    "ALTER TABLE custom_columns ADD COLUMN version_id INT NULL AFTER project_id"
                ))
            except Exception:
                pass  # 字段已存在

        db.session.expire_all()
        versions_by_project = {}
        for version in Version.query.order_by(Version.project_id, Version.sort_order, Version.id).all():
            versions_by_project.setdefault(version.project_id, []).append(version)

        for project_id, versions in versions_by_project.items():
            legacy = CustomColumn.query.filter_by(
                project_id=project_id, version_id=None
            ).order_by(CustomColumn.sort_order.asc(), CustomColumn.id.asc()).all()
            if not legacy:
                continue

            # 导入 CAN 表格前，旧版本没有自定义字段；当前遗留列中带有
            # 自定义列时，将其保留给拥有自定义数据的版本，其他版本恢复系统列。
            custom_case_counts = {
                version.id: TestCase.query.filter(
                    TestCase.project_id == project_id,
                    TestCase.version_id == version.id,
                    TestCase.custom_fields.isnot(None),
                    TestCase.custom_fields != '{}',
                ).count()
                for version in versions
            }
            imported_version = max(
                versions,
                key=lambda version: custom_case_counts.get(version.id, 0)
            ) if versions else None
            has_custom_legacy = any(not column.is_system for column in legacy)

            if imported_version and has_custom_legacy and custom_case_counts.get(imported_version.id, 0):
                # 当前数据库中这批遗留列就是 CAN 导入产生的列配置。
                for column in legacy:
                    column.version_id = imported_version.id
                for version in versions:
                    if version.id == imported_version.id:
                        continue
                    for index, system in enumerate(SYSTEM_COLUMNS):
                        db.session.add(CustomColumn(
                            project_id=project_id,
                            version_id=version.id,
                            name=system['name'],
                            key=system['key'],
                            is_system=True,
                            is_visible=True,
                            width=system['width'],
                            sort_order=index,
                            text_align='left',
                        ))
            else:
                # 其他历史项目原本所有版本共用一套列配置，先保留给第一个版本，
                # 再为其余版本复制一份，确保后续设置互不影响。
                first_version = versions[0]
                for column in legacy:
                    column.version_id = first_version.id
                for version in versions[1:]:
                    for source in legacy:
                        db.session.add(CustomColumn(
                            project_id=project_id,
                            version_id=version.id,
                            name=source.name,
                            key=source.key,
                            is_system=source.is_system,
                            is_visible=source.is_visible,
                            width=source.width,
                            sort_order=source.sort_order,
                            text_align=source.text_align or 'left',
                        ))
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        print('version_columns 迁移提示:', exc)


if __name__ == '__main__':
    with app.app_context():
        create_database()
        initialize_auth_data()
        migrate_sort_order()
    app.run(host='0.0.0.0', port=5005, debug=True)
